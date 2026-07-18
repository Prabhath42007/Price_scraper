""""
PRICE MONITORING SYSTEM
AI-Powered competitor monitoring and pricing intelligence system
"""

import json
import time
import logging
import random
from pathlib import Path
from browser import get_page
import pandas as pd
from db import save_to_db, detect_drops
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image, ImageDraw, ImageFont
 
 
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")   # price dropped
ORANGE_FILL = PatternFill("solid", fgColor="FFE5B4")   # price up (competitor raised)
GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")   # first scan / no change
 
 
def load_products() -> dict:
    with open("prices.json", "r") as f:
        return json.load(f)
 
def clean_price(raw: str) -> float | None:
    """
    Strips ₹, commas, spaces and converts to float.
    Returns None if conversion fails.
    """
    try:
        cleaned = (
            str(raw)
            .replace("₹", "")
            .replace(",", "")
            .replace("\xa0", "")   # non-breaking space
            .replace(" ", "")
            .strip()
        )
        if "-" in cleaned:
            cleaned = cleaned.split("-")[0].strip()
        return float(cleaned) if cleaned else None
    except (ValueError, TypeError):
        return None
 
def clean_results(data: list[dict]) -> list[dict]:
    """
    Clean price and rating fields IN PLACE on the original list
    so detect_drops receives proper floats, not raw strings.
    """
    cleaned = []
    for item in data:
        item["Price"]  = clean_price(item.get("Price",  ""))
        if item["Price"] is None:        # skip rows where price failed
            continue
        cleaned.append(item)
    return cleaned
 
def scrape_product(site: str, url: str, con: str, keys: dict, page: Page) -> list[dict]:
    page.goto(url, wait_until="domcontentloaded", timeout=90000)
    page.wait_for_selector(con, timeout=30000)
    page.route(
    "**/*",
    lambda route:
        route.abort()
        if route.request.resource_type
        in ["image", "font", "media"]
        else route.continue_()
    )
    cards = page.locator(con)
    count = cards.count()
    results = []
 
    for i in range(count):
        card = cards.nth(i)
        item = {}
        page.mouse.wheel(0, random.randint(30, 60))
        page.mouse.move(random.randint(100, 500), random.randint(100, 500))
        for field, selector in keys.items():
            if not selector:
                continue
            loc = card.locator(selector).first
            item[field] = loc.inner_text().strip() if loc.count() else ""
            page.mouse.wheel(0, random.randint(30, 75))
 
        if not item.get("Product Name"):
            continue
        item["Source"]     = site
        item["Time stamp"] = time.strftime("%Y-%m-%d %H:%M:%S")
        results.append(item) 
    return results

def save_data(data: list[dict], output_path: Path) -> None:
    if not data:
        return

    df = pd.DataFrame(data)
    df = df.reindex(columns=data[0].keys())
    df.drop_duplicates(subset=["Product Name", "Price"], inplace=True)
    df.dropna(how="all", inplace=True)

    # ── Capture alert values BEFORE dropping the column ──────────────────
    # We need them for row colouring but don't want them in the sheet
    alert_values = df["Drop Alert"].tolist() if "Drop Alert" in df.columns else []
    drop_pct_col = df["Drop %"].copy()      if "Drop %"    in df.columns else None

    # Fix negative drop % display (PRICE UP rows)
    if drop_pct_col is not None:
        df["Drop %"] = df["Drop %"].apply(
            lambda x: abs(x) if isinstance(x, float) and x < 0 else x
        )

    # Fix 0.0 → None so it shows blank in Excel instead of 0
    if "Drop %" in df.columns:
        df["Drop %"] = df["Drop %"].apply(
            lambda x: None if x == 0.0 else x
        )

    # Drop the alert column — colours will handle it visually
    if "Drop Alert" in df.columns:
        df.drop(columns=["Drop Alert"], inplace=True)

    sheet_name  = time.strftime("report-%Y-%m-%d_%H-%M")
    file_exists = output_path.exists()

    writer_kwargs = dict(engine="openpyxl")
    if file_exists:
        writer_kwargs.update(mode="a", if_sheet_exists="new")

    with pd.ExcelWriter(output_path, **writer_kwargs) as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]
        headers   = [c.value for c in worksheet[1]]

        def col_idx(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        price_col    = col_idx("Price")
        drop_pct_idx = col_idx("Drop %")

        # ── Header formatting ─────────────────────────────────────────────
        for cell in worksheet[1]:
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Row highlighting using saved alert_values ─────────────────────
        for row_idx, row in enumerate(
            worksheet.iter_rows(min_row=2), start=0
        ):
            if row_idx >= len(alert_values):
                break

            alert = str(alert_values[row_idx] or "")

            if alert == "URGENT":
                fill = PatternFill("solid", fgColor="FFCCCC")   # strong red
            elif alert == "YES":
                fill = PatternFill("solid", fgColor="FFAAAA")   # light red
            elif alert == "PRICE UP":
                fill = PatternFill("solid", fgColor="FFE5B4")   # orange
            else:
                continue  # First scan / blank — no fill

            for cell in row:
                cell.fill = fill

        # ── Price column format ───────────────────────────────────────────
        if price_col:
            for cell in worksheet.iter_rows(
                min_row=2, min_col=price_col, max_col=price_col
            ):
                cell[0].number_format = "₹#,##0.00"

        # ── Auto column width ─────────────────────────────────────────────
        for column in worksheet.columns:
            max_len = max(
                (len(str(c.value)) for c in column if c.value), default=10
            )
            worksheet.column_dimensions[
                column[0].column_letter
            ].width = min(max_len + 4, 60)

        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes    = "A2"

def generate_dashboard(
    products: int,
    drops: int,
    urgent: int,
    biggest_drop: float,
    output_file: str = "dashboard.png"
):
    WIDTH = 1600
    HEIGHT = 900

    img = Image.new("RGB", (WIDTH, HEIGHT), (245, 247, 250))
    draw = ImageDraw.Draw(img)

    try:
        title_font = ImageFont.truetype("arial.ttf", 52)
        card_title = ImageFont.truetype("arial.ttf", 28)
        card_value = ImageFont.truetype("arial.ttf", 54)
        body_font = ImageFont.truetype("arial.ttf", 30)
    except:
        title_font = ImageFont.load_default()
        card_title = ImageFont.load_default()
        card_value = ImageFont.load_default()
        body_font = ImageFont.load_default()

    # Header
    draw.text(
        (60, 40),
        "AI-Powered Price Monitoring Report",
        fill=(20, 20, 20),
        font=title_font
    )

    draw.text(
        (60, 110),
        "Competitor Intelligence Dashboard",
        fill=(90, 90, 90),
        font=body_font
    )

    cards = [
        ("Products Tracked", str(products)),
        ("Price Drops", str(drops)),
        ("Urgent Alerts", str(urgent)),
        ("Biggest Drop", f"{biggest_drop:.1f}%")
    ]

    start_x = 60
    start_y = 200

    card_w = 340
    card_h = 180
    gap = 30

    for i, (title, value) in enumerate(cards):
        x = start_x + i * (card_w + gap)
        y = start_y

        draw.rounded_rectangle(
            (x, y, x + card_w, y + card_h),
            radius=18,
            fill=(255, 255, 255),
            outline=(220, 220, 220),
            width=2
        )

        draw.text(
            (x + 25, y + 25),
            title,
            fill=(90, 90, 90),
            font=card_title
        )

        draw.text(
            (x + 25, y + 80),
            value,
            fill=(20, 20, 20),
            font=card_value
        )

    # Insight section
    draw.rounded_rectangle(
        (60, 450, 1540, 820),
        radius=18,
        fill=(255, 255, 255),
        outline=(220, 220, 220),
        width=2
    )

    draw.text(
        (90, 490),
        "Key Insights",
        fill=(20, 20, 20),
        font=title_font
    )

    insights = [
        f"• {drops} products experienced price drops",
        f"• {urgent} products require immediate attention",
        f"• Largest detected drop: {biggest_drop:.1f}%",
        "• Automated monitoring active across marketplaces",
        "• Historical tracking available in Excel reports"
    ]

    y = 580
    for line in insights:
        draw.text(
            (110, y),
            line,
            fill=(50, 50, 50),
            font=body_font
        )
        y += 55

    img.save(output_file, quality=95)
    

def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    logger = logging.getLogger()
 
    products = load_products()
    playwright, browser, context = get_page()
 
    start, end = 1, 5   # pages to scrape per keyword
    context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver',
                      {get: () => undefined});  
                Object.defineProperty(navigator, 'platform',
                      {get: () => 'Win32'});
                Object.defineProperty(navigator, 'languages',
                      {get: () => ['en-IN', 'en']});
            """) 
    for site in products.keys():
        for name in products[site]["pages"].keys(): 
            res = []
            page = context.new_page()
            logger.info(f"Scraping {name} from {site}...")
            for i in range(start, end + 1):
                for attempt in range(2):
                    try:
                        page_url = products[site]["pages"][name] + str(i)
                        page_res = scrape_product( site, page_url, products[site]["container"], products[site]["selectors"], page)
                        res.extend(page_res)
                        time.sleep(random.uniform(0.5, 1.5))
                        break   
                    except PlaywrightTimeoutError as e:
                        logger.warning(f"Timeout attempt {attempt+1} for {name}-{i}: {str(e)}")
                    except Exception as e:
                        logger.error(f"Error attempt {attempt+1} for {name}-{i}: {str(e)}")
                    time.sleep(random.uniform(0.5, 1))
                    page.reload()
                else:
                    logger.error(f"Failed to scrape {name}-{i} from {site}")
            page.close() 
            if not res:
                logger.error(f"\nNo results for {name} on {site}")
                continue
            res = clean_results(res) 
            res = detect_drops(name, res)
            drop_count = sum(1 for r in res if r.get("Drop Alert") in ("YES", "URGENT"))
            if drop_count:
                logger.info(f"⚠️ {drop_count} price drops detected for {name}")

            drops = sum(
                1
                for r in res
                if r.get("Drop Alert") in ("YES", "URGENT")
            )

            urgent = sum(
                1
                for r in res
                if r.get("Drop Alert") == "URGENT"
            )

            biggest_drop = max(
                (
                    r.get("Drop %", 0)
                    for r in res
                    if isinstance(r.get("Drop %"), (int, float))
                ),
                default=0
            )

            generate_dashboard(
                len(res),
                drops,
                urgent,
                biggest_drop,
                f"{name}_dashboard.png"
            )
 
            save_data(res,Path(r'c:/Python_projects/price_sheets')/f"{name}_report.xlsx")
 
            save_to_db(name, res)
 
            logger.info(f"✅ {len(res)} products saved for {name} from {end} pages on {site}")
 
    browser.close()
    playwright.stop()
 
